import sqlite3
import pandas as pd
from typing import Dict, List
from collections import Counter, defaultdict
from pathlib import Path


class AnnotationAnalyzer:
    """标注数据分析模块"""

    def __init__(self, db_path: str = 'frankenstein_annotation.db'):
        """初始化分析器"""
        if not Path(db_path).exists():
            print(f"❌ 数据库不存在: {db_path}")
            print("请先运行主标注脚本生成数据库")
            return

        self.db_path = db_path
        self.conn = sqlite3.connect(db_path)
        self.df = None
        self._load_data()

    def _load_data(self):
        """加载数据"""
        try:
            self.df = pd.read_sql_query('SELECT * FROM annotations', self.conn)
            print(f"✓ 成功加载 {len(self.df)} 条标注记录")
        except Exception as e:
            print(f"❌ 加载数据失败: {e}")

    def print_summary(self):
        """打印汇总信息"""
        if self.df is None or len(self.df) == 0:
            print("❌ 无数据可分析")
            return

        print("\n" + "=" * 70)
        print("《弗兰肯斯坦》标注语料库分析报告")
        print("=" * 70)

        # 基本统计
        print(f"\n【基本统计】")
        print(f"总标注数:     {len(self.df)}")
        print(f"涉及章节:     {self.df['chapter_num'].nunique()}")
        print(f"命名冲突:     {len(self.df[self.df['naming_conflict'] == '是'])}")
        print(f"标注日期:     {self.df['timestamp'].max() if 'timestamp' in self.df.columns else '未记录'}")

        # 叙事者分析
        print(f"\n【叙事者分布】")
        narrator_counts = self.df['narrator'].value_counts()
        for narrator, count in narrator_counts.items():
            percentage = count / len(self.df) * 100
            print(f"  {narrator:12} : {count:3} 条 ({percentage:5.1f}%)")

        # 称呼分析
        print(f"\n【称呼频率分析】")
        appellation_counts = self.df['appellation'].value_counts()
        for appellation, count in appellation_counts.items():
            percentage = count / len(self.df) * 100
            print(f"  {appellation:12} : {count:3} 条 ({percentage:5.1f}%)")

        # 行为分析
        print(f"\n【伴随行为分析】")
        behaviors = defaultdict(int)
        for behavior_str in self.df['behavior']:
            if pd.isna(behavior_str) or behavior_str == '':
                continue
            behavior_list = str(behavior_str).split(',')
            for behavior in behavior_list:
                behavior = behavior.strip()
                if behavior and behavior != '未标注':
                    behaviors[behavior] += 1

        for behavior, count in sorted(behaviors.items(), key=lambda x: x[1], reverse=True):
            print(f"  {behavior:12} : {count:3} 次")

        # 命名冲突分析
        conflicts = self.df[self.df['naming_conflict'] == '是']
        if len(conflicts) > 0:
            print(f"\n【命名冲突分析】(共 {len(conflicts)} 处)")

            conflict_types = defaultdict(int)
            for note in conflicts['quality_note']:
                if pd.notna(note) and ':' in str(note):
                    conflict_type = str(note).split(':')[0]
                    conflict_types[conflict_type] += 1

            for ctype, count in sorted(conflict_types.items(), key=lambda x: x[1], reverse=True):
                print(f"  {ctype:12} : {count:3} 处")

        print("\n" + "=" * 70 + "\n")

    def export_to_excel(self, output_file: str = 'analysis_report.xlsx'):
        """导出完整分析报告"""
        if self.df is None or len(self.df) == 0:
            print("❌ 无数据可导出")
            return

        print(f"导出分析报告: {output_file}...", end=' ')

        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Sheet 1: 原始标注
                self.df.to_excel(writer, sheet_name='原始标注', index=False)

                # Sheet 2: 叙事者统计
                narrator_stats = self.df['narrator'].value_counts().reset_index()
                narrator_stats.columns = ['叙事者', '出现次数']
                narrator_stats['百分比'] = (narrator_stats['出现次数'] / narrator_stats['出现次数'].sum() * 100).round(2)
                narrator_stats.to_excel(writer, sheet_name='叙事者统计', index=False)

                # Sheet 3: 称呼频率
                appellation_stats = self.df['appellation'].value_counts().reset_index()
                appellation_stats.columns = ['称呼', '出现次数']
                appellation_stats['百分比'] = (appellation_stats['出现次数'] / appellation_stats['出现次数'].sum() * 100).round(2)
                appellation_stats.to_excel(writer, sheet_name='称呼频率', index=False)

                # Sheet 4: 命名冲突
                conflicts = self.df[self.df['naming_conflict'] == '是']
                if len(conflicts) > 0:
                    conflicts[[
                        'chapter_num', 'narrator', 'appellation', 'context', 'quality_note'
                    ]].to_excel(writer, sheet_name='命名冲突', index=False)

                # Sheet 5: 章节统计
                chapter_stats = self.df['chapter_num'].value_counts().reset_index()
                chapter_stats.columns = ['章节', '标注数']
                chapter_stats.to_excel(writer, sheet_name='章节统计', index=False)

                # 格式化
                from openpyxl.styles import PatternFill, Font, Alignment

                workbook = writer.book
                for sheet_name in workbook.sheetnames:
                    worksheet = writer.sheets[sheet_name]

                    # 列头格式
                    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    header_font = Font(bold=True, color='FFFFFF')

                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', wrap_text=True)

                    # 自动调整列宽
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

            print("✓")
            return output_file

        except Exception as e:
            print(f"❌ {e}")

    def show_sample_conflicts(self, limit: int = 5):
        """显示样本命名冲突"""
        if self.df is None or len(self.df) == 0:
            print("❌ 无数据")
            return

        conflicts = self.df[self.df['naming_conflict'] == '是']

        if len(conflicts) == 0:
            print("未发现命名冲突")
            return

        print(f"\n【命名冲突样本】(共 {len(conflicts)} 处，显示前 {min(limit, len(conflicts))} 处)")
        print("-" * 70)

        for idx, (_, row) in enumerate(conflicts.head(limit).iterrows(), 1):
            print(f"\n{idx}. {row['chapter_num']}")
            print(f"   叙事者: {row['narrator']}")
            print(f"   称呼:   {row['appellation']}")
            print(f"   分析:   {row['quality_note']}")
            print(f"   语境:   ...{row['context'][:100]}...")

        print("\n" + "-" * 70)

    def get_narrator_appellation_matrix(self) -> pd.DataFrame:
        """生成叙事者-称呼矩阵"""
        if self.df is None or len(self.df) == 0:
            print("❌ 无数据")
            return pd.DataFrame()

        matrix = pd.crosstab(self.df['narrator'], self.df['appellation'])
        return matrix

    def print_narrator_preferences(self):
        """打印各叙事者的称呼偏好"""
        if self.df is None or len(self.df) == 0:
            print("❌ 无数据")
            return

        print("\n【各叙事者的称呼偏好】")
        print("-" * 50)

        for narrator in self.df['narrator'].unique():
            if narrator == '未知':
                continue

            narrator_df = self.df[self.df['narrator'] == narrator]
            appellation_counts = narrator_df['appellation'].value_counts()

            print(f"\n{narrator}:")
            for appellation, count in appellation_counts.head(3).items():
                print(f"  • {appellation}: {count} 次")

    def print_narrator_behavior_matrix(self):
        """打印叙事者-行为矩阵"""
        if self.df is None or len(self.df) == 0:
            print("❌ 无数据")
            return

        print("\n【各叙事者的主要行为】")
        print("-" * 50)

        for narrator in self.df['narrator'].unique():
            if narrator == '未知':
                continue

            narrator_df = self.df[self.df['narrator'] == narrator]

            behaviors = defaultdict(int)
            for behavior_str in narrator_df['behavior']:
                if pd.isna(behavior_str) or behavior_str == '':
                    continue
                behavior_list = str(behavior_str).split(',')
                for behavior in behavior_list:
                    behavior = behavior.strip()
                    if behavior and behavior != '未标注':
                        behaviors[behavior] += 1

            if behaviors:
                print(f"\n{narrator}:")
                for behavior, count in sorted(behaviors.items(), key=lambda x: x[1], reverse=True)[:3]:
                    print(f"  • {behavior}: {count} 次")

    def close(self):
        """关闭数据库连接"""
        if self.conn:
            self.conn.close()


class TextValidator:
    """文本验证模块"""

    @staticmethod
    def check_file(file_path: str) -> Dict:
        """检查文本文件"""
        result = {
            'exists': False,
            'readable': False,
            'size': 0,
            'encoding': 'unknown',
            'has_content': False
        }

        try:
            path = Path(file_path)
            result['exists'] = path.exists()

            if result['exists']:
                result['size'] = path.stat().st_size

                # 尝试读取
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                        result['readable'] = True
                        result['encoding'] = 'UTF-8'
                        result['has_content'] = len(content) > 0
                except:
                    result['readable'] = False

        except Exception as e:
            print(f"检查失败: {e}")

        return result


# ============ 主程序 ============

if __name__ == '__main__':
    # 检查主标注脚本是否已运行
    print("检查标注数据...")

    if not Path('frankenstein_annotation.db').exists():
        print("\n❌ 数据库不存在！")
        print("请先运行主标注脚本:")
        print("  python annotation_pipeline.py")
        exit(1)

    # 创建分析器
    analyzer = AnnotationAnalyzer()

    if analyzer.df is None or len(analyzer.df) == 0:
        print("❌ 无法加载标注数据")
        exit(1)

    # 打印汇总
    analyzer.print_summary()

    # 打印偏好分析
    analyzer.print_narrator_preferences()

    # 打印行为分析
    analyzer.print_narrator_behavior_matrix()

    # 显示样本冲突
    analyzer.show_sample_conflicts(limit=5)

    # 导出报告
    analyzer.export_to_excel('analysis_report.xlsx')

    # 关闭
    analyzer.close()

    print("\n✓ 分析完成！")